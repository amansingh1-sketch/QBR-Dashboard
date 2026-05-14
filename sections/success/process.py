#!/usr/bin/env python3
"""Regenerate Customer Success section JSON files.

Reads:
  - ~/Downloads/[Customer Success] Overall Key Metrics.xlsx
  - ~/Downloads/[Growth Retention] Key Metrics.xlsx
  - ~/Downloads/[Strategic CS] Key Metrics.xlsx
  - /tmp/qbr-raw/cs/{created_q1,closed_q1,open_q2}.json  (HubSpot MCP cache)

Writes:
  - data/success/mrr-change-q1-fy2026.json
  - data/success/success-metrics-q1-fy2026.json
  - data/success/product-adoption-q1-fy2026.json
  - data/success/expansion-pipeline-q1-fy2026.json
"""
import warnings; warnings.filterwarnings("ignore")
import json, os, sys
from pathlib import Path
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
OUT_DIR = REPO_ROOT / "data" / "success"
OUT_DIR.mkdir(parents=True, exist_ok=True)

SRC = {
    "total":     str(Path.home() / "Downloads/[Customer Success] Overall Key Metrics.xlsx"),
    "scaled":    str(Path.home() / "Downloads/[Growth Retention] Key Metrics.xlsx"),
    "strategic": str(Path.home() / "Downloads/[Strategic CS] Key Metrics.xlsx"),
}

# ---------- 1. MRR Change ----------

ROW_LABELS = {
    "Starting MRR": "starting_mrr",
    "New": "new",
    "Expansion": "expansion",
    "Downgrade": "downgrade",
    "Churn": "churn",
    "Reactivation": "reactivation",
    "Closing MRR": "closing_mrr",
    "Net Expansions": "net_expansions",
    "Net Expansion as % of start": "net_expansion_pct",
    "Expansion as % of start": "expansion_pct",
    "Downgrade as % of start": "downgrade_pct",
    "Churn as % of start": "churn_pct",
}
SECTION_LABELS = {
    "New customers post Feb 2026": "new_customers",
    "Existing customer base post Feb 2026": "existing_customers",
    "Total": "total",
}


def parse_mrr_change(path: str) -> dict:
    wb = load_workbook(path, data_only=True)
    ws = wb["MRR Change"]
    out = {"new_customers": None, "existing_customers": None, "total": None}
    # First pass: locate every section header so we know each section's bounds.
    section_rows = []
    for i in range(1, ws.max_row + 1):
        a = ws.cell(row=i, column=1).value
        if a in SECTION_LABELS:
            section_rows.append((i, SECTION_LABELS[a]))
    # Second pass: read each section, scanning only up to the next header.
    for idx, (start, key) in enumerate(section_rows):
        end = section_rows[idx + 1][0] if idx + 1 < len(section_rows) else ws.max_row + 1
        section = {"months": ["Feb 2026", "Mar 2026", "Apr 2026", "May 2026"]}
        for j in range(start, end):
            label = ws.cell(row=j, column=1).value
            if label in ROW_LABELS:
                vals = [ws.cell(row=j, column=c).value for c in range(3, 7)]
                section[ROW_LABELS[label]] = [round(v, 2) if isinstance(v, (int, float)) else None for v in vals]
        out[key] = section
    return out

# ---------- 2. Customer Movements ----------

def parse_movements() -> dict:
    wb = load_workbook(SRC["total"], data_only=True)
    ws = wb["Extract - Movt"]
    months = ["2026-02-01", "2026-03-01", "2026-04-01"]  # Q1
    init = lambda: {m: 0 for m in months}
    plan_upgrades, m2a, a2m, aiva_act, aiva_deact = init(), init(), init(), init(), init()
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] != "Month" or r[1] is None:
            continue
        m = r[1].strftime("%Y-%m-01")
        if m not in months:
            continue
        users = int(r[6] or 0)
        if r[5] == "Plan Upgrade": plan_upgrades[m] += users
        if r[3] == "Monthly→Annual": m2a[m] += users
        if r[3] == "Annual→Monthly": a2m[m] += users
        if r[4] == "Inactive→Active": aiva_act[m] += users
        if r[4] == "Active→Inactive": aiva_deact[m] += users
    return {
        "months": months,
        "plan_upgrades": [plan_upgrades[m] for m in months],
        "m2a": [m2a[m] for m in months],
        "a2m": [a2m[m] for m in months],
        "aiva_activations": [aiva_act[m] for m in months],
        "aiva_deactivations": [aiva_deact[m] for m in months],
    }

# ---------- 3. Product Adoption ----------

def parse_product_adoption() -> dict:
    """
    Layout in the xlsx:
      - 3 sections stacked vertically, each headed by a value in col B:
        "New customers..."   (~row 2)
        "Existing customers..." (~row 42)
        "Total Customer base"   (~row 82)
      - Within a section: monthly columns at cols 5..9 (Jan/Feb/Mar/Apr/May).
        Module group label appears in col B only on the FIRST row of each group;
        subsequent rows in the same group have B = None.
        Metric label is in col C.
    """
    wb = load_workbook(SRC["scaled"], data_only=True)
    ws = wb["Product adoption"]
    months_full = ["Jan 2026", "Feb 2026", "Mar 2026", "Apr 2026", "May 2026"]
    # Q1 view = Feb/Mar/Apr only.
    months_keep = ["Feb 2026", "Mar 2026", "Apr 2026"]
    keep_idx = [months_full.index(m) for m in months_keep]
    months = months_keep

    SECTION_KEYS = {
        "new":      lambda b: isinstance(b, str) and "New customers" in b,
        "existing": lambda b: isinstance(b, str) and "Existing customers" in b,
        "total":    lambda b: isinstance(b, str) and b.strip() == "Total Customer base",
    }
    CUSTOMER_BASE_LABELS = {"New Customer base", "Existing Customer base", "Total Customer base"}

    cohorts: dict[str, list] = {"new": [], "existing": [], "total": []}
    current = None
    current_module = None  # forward-fill module group within a section

    for i in range(1, min(ws.max_row + 1, 200)):
        b = ws.cell(row=i, column=2).value
        c = ws.cell(row=i, column=3).value

        # Section header detection (in order of likelihood).
        new_section = None
        for key, pred in SECTION_KEYS.items():
            if pred(b):
                new_section = key; break
        if new_section is not None:
            current = new_section
            current_module = None
            continue

        if current is None:
            continue

        # A row that introduces a new module group has B set; reset module then.
        if isinstance(b, str) and b.strip():
            current_module = b.strip()

        label = c if isinstance(c, str) else None
        if not label:
            continue

        vals = [ws.cell(row=i, column=col).value for col in range(5, 10)]
        if all(v is None for v in vals):
            continue
        if not all(isinstance(v, (int, float, type(None))) for v in vals):
            continue

        # The "Customer base" line is a section summary, not a module metric.
        is_customer_base = label in CUSTOMER_BASE_LABELS
        kept = [vals[k] for k in keep_idx]
        cohorts[current].append({
            "module_group": None if is_customer_base else current_module,
            "metric": label,
            "values": [round(v) if isinstance(v, (int, float)) else None for v in kept],
            "is_customer_base": is_customer_base,
        })

    return {"months": months, "cohorts": cohorts}

# ---------- 4 & 5. Expansion pipeline ----------

STAGE_LABELS = {
    "ffb2325a-8b75-4c1a-8caa-5b32fa6adf91": "CS Identified Lead",
    "3b7a1229-e60d-4313-aa1f-5d4ceb839e8b": "Deal Qualification",
    "90ee7819-5832-49f1-a5c0-edcb67ee9dd1": "Demo",
    "ec3848a9-d1ae-4fc2-97a6-a21ba85d7e6c": "Proposal",
    "b6e6b756-f51a-45ef-a862-8a40389f896c": "Closed Won",
    "76d5f953-79fe-4be1-93de-b70efac45b87": "Closed Lost",
}
AIVA_SUBTYPES = {"Add On - Voice Agent", "Add On - Outbound Voice Agent"}


def group_by_stage(deals: list) -> list:
    by_stage: dict[str, dict] = {}
    for d in deals:
        stage = d.get("dealstage") or "unknown"
        label = STAGE_LABELS.get(stage, stage[:8])
        amount = float(d.get("amount") or 0)
        b = by_stage.setdefault(label, {"deals": 0, "amount": 0.0})
        b["deals"] += 1
        b["amount"] += amount
    return sorted(
        [{"stage": k, "deals": v["deals"], "amount": round(v["amount"], 2)} for k, v in by_stage.items()],
        key=lambda x: -x["deals"],
    )


def cohort_summary(deals: list) -> dict:
    return {
        "total": len(deals),
        "totalAmount": round(sum(float(d.get("amount") or 0) for d in deals), 2),
        "byStage": group_by_stage(deals),
    }


def parse_expansion_pipeline() -> dict:
    cs_dir = Path("/tmp/qbr-raw/cs")
    cohorts = {}
    for name in ("created_q1", "closed_q1", "open_q2"):
        with open(cs_dir / f"{name}.json") as f:
            cohorts[name] = json.load(f)
    aiva = lambda ds: [d for d in ds if d.get("opportunity_sub_type") in AIVA_SUBTYPES]
    return {
        "all": {
            "createdInQ1": cohort_summary(cohorts["created_q1"]),
            "closedInQ1": cohort_summary(cohorts["closed_q1"]),
            "openIntoQ2": cohort_summary(cohorts["open_q2"]),
        },
        "aiva": {
            "createdInQ1": cohort_summary(aiva(cohorts["created_q1"])),
            "closedInQ1": cohort_summary(aiva(cohorts["closed_q1"])),
            "openIntoQ2": cohort_summary(aiva(cohorts["open_q2"])),
        },
    }


def write_json(name: str, obj) -> None:
    path = OUT_DIR / f"{name}-q1-fy2026.json"
    with open(path, "w") as f:
        json.dump(obj, f, indent=2)
    print(f"wrote {path.relative_to(REPO_ROOT)}")


def main() -> None:
    write_json("mrr-change", {seg: parse_mrr_change(p) for seg, p in SRC.items()})
    write_json("success-metrics", parse_movements())
    write_json("product-adoption", parse_product_adoption())
    write_json("expansion-pipeline", parse_expansion_pipeline())
    print("Customer Success: all 4 files written.")


if __name__ == "__main__":
    main()
